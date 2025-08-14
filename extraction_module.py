import pandas as pd
from openpyxl import load_workbook, Workbook

def excel_to_csv_from_middle(excel_path, csv_path, sheet_name=0, start_row=9, usecols=None):
    try:
        df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name,
            skiprows=start_row,
            usecols=usecols
        )
        # df.dropna(how='all', inplace=True)
        df.to_csv(csv_path, index=False)
        print(f"Successfully converted Excel to CSV and saved at '{csv_path}'.")
    except Exception as e:
        print(f"Error occured while converting Excel to CSV: {e}")

def csv_to_excel_in_middle(csv_path, excel_path, sheet_name="Sheet1", start_row=9, start_col=4):
    try:
        df = pd.read_csv(csv_path)

        try:
            wb = load_workbook(excel_path)
        except FileNotFoundError:
            wb = Workbook()

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        for j, col_name in enumerate(df.columns):
            ws.cell(row=start_row, column=start_col + j + 1, value=col_name)

        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i + 1, column=start_col + j + 1, value=value)

        wb.save(excel_path)
        print(f"Successfully converted CSV to Excel and saved at:'{excel_path}', starting at row {start_row + 1}, column {start_col + 1}.")
    except Exception as e:
        print(f"Error occurred while converting CSV to the Excel file: {e}")

def excel_to_csv_detect_table(excel_path, csv_path, sheet_name=0):
    try:
        df_all = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, engine='openpyxl')

        header_row = None
        for idx, row in df_all.iterrows():
            if row.notna().sum() > 2:
                header_row = idx
                break

        if header_row is None:
            print("No table was found in the Excel sheet.")
            return

        df_table = pd.read_excel(excel_path, sheet_name=sheet_name, skiprows=header_row, engine='openpyxl')
        df_table.dropna(how='all', inplace=True)
        df_table.to_csv(csv_path, index=False)
        print(f"A table starting at row {header_row + 1} was found.")
        print(f"Successfully converted Excel to CSV and saved at: '{csv_path}'")
    except Exception as e:
        print(f"Error occurred while converting Excel to CSV file: {e}")

def excel_to_csv_full(excel_path, csv_path, sheet_name=0):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
        df.to_csv(csv_path, index=False)
        print(f"Successfully converted Excel to CSV and saved at: '{csv_path}'")
    except Exception as e:
        print(f"Error occurred while converting Excel to CSV file: {e}")

def csv_to_excel_full(csv_path, excel_path, sheet_name="Sheet1"):
    try:
        df = pd.read_csv(csv_path)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        ws.append(list(df.columns))  # headers
        for row in df.values:
            ws.append(list(row))

        wb.save(excel_path)
        print(f"Successfully converted CSV to Excel:'{excel_path}'.")
    except Exception as e:
        print(f"Error occurred while converting CSV to the Excel file: {e}")

def main():
    action = input(
        "\nWhat do you want to do?\n"
        "- Type 'excel to csv middle' to convert a table from inside an Excel sheet.\n"
        "- Type 'csv to excel middle' to write a CSV into a specific cell in Excel.\n"
        "- Type 'excel to csv auto' to detect a table and convert it.\n"
        "- Type 'excel to csv' to convert the whole Excel sheet to CSV.\n"
        "- Type 'csv to excel' to convert a whole CSV file to Excel.\n\n"
        "Please enter the choice: "
    ).strip().lower()

    if action == "excel to csv middle":
        excel_path = input("Excel file path: ")
        csv_path = input("CSV output path: ")
        row = int(input("Starting row number of the table (e.g., 10): ")) - 1
        cols = input("Columns to read (e.g., 'E:J') or leave blank: ").strip()
        usecols = cols if cols else None
        excel_to_csv_from_middle(excel_path, csv_path, start_row=row, usecols=usecols)

    elif action == "csv to excel middle":
        csv_path = input("CSV file path: ")
        excel_path = input("Excel file path to write to (will be created if it doesn't exist): ")
        row = int(input("Row number to start writing in Excel (e.g., 10): ")) - 1
        col = int(input("Column number to start writing (A=1, B=2, etc.): ")) - 1
        csv_to_excel_in_middle(csv_path, excel_path, start_row=row, start_col=col)

    elif action == "excel to csv auto":
        excel_path = input("Excel file path: ")
        csv_path = input("CSV output path: ")
        excel_to_csv_detect_table(excel_path, csv_path)

    elif action == "excel to csv":
        excel_path = input("Excel file path: ")
        csv_path = input("CSV output path: ")
        excel_to_csv_full(excel_path, csv_path)

    elif action == "csv to excel":
        csv_path = input("CSV file path: ")
        excel_path = input("Excel file output path: ")
        csv_to_excel_full(csv_path, excel_path)

    else:
        print("Sorry, I didn’t understand that. Please type one of the listed commands exactly.")

if __name__ == "__main__":
    main()
